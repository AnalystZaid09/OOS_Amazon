# app.py - Streamlit-ready (Template-first pivot; fallback workbook)
import streamlit as st
import pandas as pd
import io
import os
import traceback
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="OOS Amazon Analysis", page_icon="📊", layout="wide")

# -------------------------
# CSS + DOC legend
# -------------------------
st.markdown(
    """
<style>
    .stApp { max-width: 100%; }
    .doc-legend { display: flex; gap: 15px; flex-wrap: wrap; margin: 20px 0; }
    .legend-item { display: flex; align-items: center; gap: 8px; }
    .legend-box { width: 30px; height: 30px; border: 1px solid #ddd; border-radius: 4px; }
</style>
""",
    unsafe_allow_html=True,
)

st.header("🎨 DOC Color Legend")
st.markdown(
    """
<div class="doc-legend">
    <div class="legend-item"><div class="legend-box" style="background-color: #8B0000;"></div><span><b>0-7 days</b> (Critical)</span></div>
    <div class="legend-item"><div class="legend-box" style="background-color: #FF8C00;"></div><span><b>7-15 days</b> (Low)</span></div>
    <div class="legend-item"><div class="legend-box" style="background-color: #006400;"></div><span><b>15-30 days</b> (Good)</span></div>
    <div class="legend-item"><div class="legend-box" style="background-color: #B8860B;"></div><span><b>30-45 days</b> (Optimal)</span></div>
    <div class="legend-item"><div class="legend-box" style="background-color: #0F52BA;"></div><span><b>45-60 days</b> (High)</span></div>
    <div class="legend-item"><div class="legend-box" style="background-color: #8B4513;"></div><span><b>60-90 days</b> (Excess)</span></div>
</div>
""",
    unsafe_allow_html=True,
)

# -------------------------
# Helpers
# -------------------------
def color_doc(val):
    try:
        doc = float(val)
        if 0 <= doc < 7:
            return "background-color: #8B0000; color: white"
        elif 7 <= doc < 15:
            return "background-color: #FF8C00; color: white"
        elif 15 <= doc < 30:
            return "background-color: #006400; color: white"
        elif 30 <= doc < 45:
            return "background-color: #B8860B; color: white"
        elif 45 <= doc < 60:
            return "background-color: #0F52BA; color: white"
        elif 60 <= doc < 90:
            return "background-color: #8B4513; color: white"
        else:
            return "background-color: #222222; color: white"
    except:
        return ""

def filter_oos(df: pd.DataFrame) -> pd.DataFrame:
    """OOS: sirf afn-fulfillable-quantity == 0"""
    col = "afn-fulfillable-quantity"
    if col not in df.columns:
        return df.iloc[0:0].copy()
    qty = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df[qty == 0].copy()

def filter_overstock(df: pd.DataFrame, threshold: float = 90.0) -> pd.DataFrame:
    """Overstock: DOC >= threshold (default 90)"""
    if "DOC" not in df.columns:
        return df.iloc[0:0].copy()
    doc = pd.to_numeric(df["DOC"], errors="coerce")
    return df[doc >= threshold].copy()

def create_excel_with_doc_format(df: pd.DataFrame) -> bytes:
    """
    Create an XLSX bytes object from df and apply DOC conditional formatting
    matching the app's color buckets. Returns bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header + rows (preserve types where possible)
    header = list(df.columns)
    ws.append(header)
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    # Find DOC column index (case-insensitive contains 'doc')
    header_row = [cell.value for cell in ws[1]]
    doc_idx = None
    for idx, h in enumerate(header_row, start=1):
        if h and "doc" in str(h).strip().lower():
            doc_idx = idx
            break

    # Color map (same as app)
    col_map = {
        "dark_red": "8B0000",
        "dark_orange": "FF8C00",
        "dark_green": "006400",
        "golden": "B8860B",
        "sky": "0F52BA",
        "saddle": "8B4513",
        "white": "FFFFFF",
    }

    if doc_idx is not None:
        try:
            last_row = ws.max_row
            col_letter = get_column_letter(doc_idx)
            first_data_row = 2
            rng = f"{col_letter}{first_data_row}:{col_letter}{last_row}"

            rules = [
                CellIsRule(operator='between', formula=['0','6.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_red"], end_color=col_map["dark_red"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['7','14.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_orange"], end_color=col_map["dark_orange"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['15','29.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_green"], end_color=col_map["dark_green"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['30','44.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["golden"], end_color=col_map["golden"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['45','59.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["sky"], end_color=col_map["sky"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['60','89.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["saddle"], end_color=col_map["saddle"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
            ]

            for rrule in rules:
                ws.conditional_formatting.add(rng, rrule)
        except Exception:
            pass

    # Auto-adjust widths
    try:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    except Exception:
        pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def fill_template_and_get_bytes(template_path: str, df: pd.DataFrame, table_name: str = "DataTable"):
    """
    Load an Excel template (xlsx/xlsm) and ensure there's an Excel Table named `table_name`.
    - If the table exists: replace header+rows and update the table.ref.
    - If the table is missing: create a new sheet named table_name or table_name_1 and add a Table.
    Additionally apply conditional formatting (CellIs rules) to the DOC column.
    Returns BytesIO of the modified workbook.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill, Font

    # keep_vba=True so macros are preserved when template is xlsm
    wb = load_workbook(template_path, keep_vba=True)
    table_sheet = None
    table_obj = None

    # 1) try to find existing table named table_name
    for ws in wb.worksheets:
        for tbl in ws._tables:
            if tbl.displayName == table_name:
                table_sheet = ws
                table_obj = tbl
                break
        if table_obj:
            break

    # helper: parse A1 addresses
    def cell_to_rowcol(cell):
        import re
        m = re.match(r"([A-Z]+)(\d+)", cell)
        if not m:
            raise RuntimeError("Unexpected table ref format")
        col_letters, row = m.groups()
        col = 0
        for ch in col_letters:
            col = col * 26 + (ord(ch) - ord('A') + 1)
        return int(row), col

    if table_obj is not None:
        # update existing table
        ref = table_obj.ref
        start_cell, end_cell = ref.split(":")
        start_row, start_col = cell_to_rowcol(start_cell)
        end_row, end_col = cell_to_rowcol(end_cell)

        # clear old rows under header
        for r in range(start_row + 1, end_row + 1):
            for c in range(start_col, end_col + 1):
                table_sheet.cell(row=r, column=c).value = None

        # write new header and rows
        header = list(df.columns)
        for idx, col_name in enumerate(header):
            table_sheet.cell(row=start_row, column=start_col + idx, value=col_name)

        for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 1):
            for c_idx, v in enumerate(row, start=start_col):
                table_sheet.cell(row=r_idx, column=c_idx, value=v)

        # update ref
        new_end_row = start_row + len(df)
        new_end_col = start_col + len(header) - 1
        table_obj.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(new_end_col)}{new_end_row}"
        target_ws = table_sheet
        header_row_idx = start_row
        first_data_row = start_row + 1
    else:
        # create a new sheet named table_name (avoid collisions)
        sheet_name = table_name
        base_name = sheet_name
        i = 1
        while sheet_name in [ws.title for ws in wb.worksheets]:
            sheet_name = f"{base_name}_{i}"
            i += 1
        ws_new = wb.create_sheet(sheet_name)

        header = list(df.columns)
        ws_new.append(header)
        for row in df.itertuples(index=False, name=None):
            ws_new.append(row)

        max_row = ws_new.max_row
        max_col = ws_new.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws_new.add_table(table)

        target_ws = ws_new
        header_row_idx = 1
        first_data_row = 2

    # --- apply conditional formatting to DOC column in target_ws ---
    # find DOC column index by header row
    header_cells = list(target_ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]
    doc_col_idx = None
    for idx, cell in enumerate(header_cells, start=1):
        if cell.value and str(cell.value).strip().lower() == "doc":
            doc_col_idx = idx
            break
    if doc_col_idx is None:
        for idx, cell in enumerate(header_cells, start=1):
            if cell.value and "doc" in str(cell.value).strip().lower():
                doc_col_idx = idx
                break

    col_map = {
        "dark_red": "8B0000",
        "dark_orange": "FF8C00",
        "dark_green": "006400",
        "golden": "B8860B",
        "sky": "0F52BA",
        "saddle": "8B4513",
        "white": "FFFFFF",
    }

    if doc_col_idx is not None:
        try:
            last_row = target_ws.max_row
            col_letter = get_column_letter(doc_col_idx)
            rng = f"{col_letter}{first_data_row}:{col_letter}{last_row}"

            # create rules for buckets using CellIsRule
            rules = [
                CellIsRule(operator='between', formula=['0','6.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_red"], end_color=col_map["dark_red"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['7','14.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_orange"], end_color=col_map["dark_orange"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['15','29.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["dark_green"], end_color=col_map["dark_green"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['30','44.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["golden"], end_color=col_map["golden"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['45','59.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["sky"], end_color=col_map["sky"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
                CellIsRule(operator='between', formula=['60','89.9999'],
                           stopIfTrue=True,
                           fill=PatternFill(start_color=col_map["saddle"], end_color=col_map["saddle"], fill_type="solid"),
                           font=Font(color="FFFFFF")),
            ]

            for rrule in rules:
                target_ws.conditional_formatting.add(rng, rrule)
        except Exception:
            pass
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def create_fallback_workbook(df: pd.DataFrame, sort_desc: bool, sheet_name: str, parent_col: str = None, selected_brands = None):
    """
    Create fallback workbook bytes (openpyxl) with:
      - data sheet (DOC styling)
      - PivotSummary (aggregated sums by Brand + parent)
      - ChartData and Chart
      - HowToPivot sheet
    """
    colors = {
        "dark_red": "8B0000",
        "dark_orange": "FF8C00",
        "dark_green": "006400",
        "golden": "B8860B",
        "sky": "0F52BA",
        "saddle": "8B4513",
        "white": "FFFFFF",
    }

    working = df.copy()
    if selected_brands:
        working = working[working["Brand"].isin(selected_brands)].copy()
    working = working.sort_values(by="DOC", ascending=(not sort_desc)).reset_index(drop=True)

    if parent_col and parent_col in working.columns:
        agg = working.groupby(
            ["Brand", parent_col], dropna=False
        )[["DOC", "DRR", "Total CP"]].sum().reset_index()
        agg["Brand_Parent"] = agg["Brand"].astype(str) + " | " + agg[parent_col].astype(str)

    elif "Brand" in working.columns:
        agg = working.groupby(
            ["Brand"], dropna=False
        )[["DOC", "DRR", "Total CP"]].sum().reset_index()
        agg["Brand_Parent"] = agg["Brand"].astype(str)

    else:
        agg = pd.DataFrame(columns=["Brand_Parent", "DOC", "DRR", "Total CP"])

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # write data
    for r in dataframe_to_rows(working, index=False, header=True):
        ws.append(r)

    # apply DOC fills
    try:
        if "DOC" in working.columns:
            doc_idx = list(working.columns).index("DOC") + 1
            from openpyxl.styles import PatternFill, Font
            def fill_for_val(v):
                try:
                    v = float(v)
                except:
                    return PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid"), Font(color="000000")
                if 0 <= v < 7:
                    return PatternFill(start_color=colors["dark_red"], end_color=colors["dark_red"], fill_type="solid"), Font(color="FFFFFF")
                if 7 <= v < 15:
                    return PatternFill(start_color=colors["dark_orange"], end_color=colors["dark_orange"], fill_type="solid"), Font(color="FFFFFF")
                if 15 <= v < 30:
                    return PatternFill(start_color=colors["dark_green"], end_color=colors["dark_green"], fill_type="solid"), Font(color="FFFFFF")
                if 30 <= v < 45:
                    return PatternFill(start_color=colors["golden"], end_color=colors["golden"], fill_type="solid"), Font(color="FFFFFF")
                if 45 <= v < 60:
                    return PatternFill(start_color=colors["sky"], end_color=colors["sky"], fill_type="solid"), Font(color="FFFFFF")
                if 60 <= v < 90:
                    return PatternFill(start_color=colors["saddle"], end_color=colors["saddle"], fill_type="solid"), Font(color="FFFFFF")
                return PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid"), Font(color="000000")

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=doc_idx)
                fill, font = fill_for_val(cell.value)
                cell.fill = fill
                cell.font = font
    except Exception:
        pass

    # add an Excel Table (DataTable)
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="DataTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
    except Exception:
        pass

    # PivotSummary
    ws_pivot = wb.create_sheet("PivotSummary")
    for r in dataframe_to_rows(agg, index=False, header=True):
        ws_pivot.append(r)

    # apply DOC fills on PivotSummary
    try:
        headers = [c.value for c in ws_pivot[1]]
        if "DOC" in headers:
            doc_idx_p = headers.index("DOC") + 1
            from openpyxl.styles import PatternFill, Font
            for row in range(2, ws_pivot.max_row + 1):
                cell = ws_pivot.cell(row=row, column=doc_idx_p)
                try:
                    val = float(cell.value)
                except:
                    val = None
                if val is None:
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    cell.font = Font(color="000000")
                elif 0 <= val < 7:
                    cell.fill = PatternFill(start_color=colors["dark_red"], end_color=colors["dark_red"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif 7 <= val < 15:
                    cell.fill = PatternFill(start_color=colors["dark_orange"], end_color=colors["dark_orange"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif 15 <= val < 30:
                    cell.fill = PatternFill(start_color=colors["dark_green"], end_color=colors["dark_green"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif 30 <= val < 45:
                    cell.fill = PatternFill(start_color=colors["golden"], end_color=colors["golden"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif 45 <= val < 60:
                    cell.fill = PatternFill(start_color=colors["sky"], end_color=colors["sky"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif 60 <= val < 90:
                    cell.fill = PatternFill(start_color=colors["saddle"], end_color=colors["saddle"], fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                else:
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    cell.font = Font(color="000000")
    except Exception:
        pass

    # ChartData + Chart
    ws_chartdata = wb.create_sheet("ChartData")
    if not agg.empty:
        for r in dataframe_to_rows(agg[["Brand_Parent", "DOC", "DRR","Total CP"]], index=False, header=True):
            ws_chartdata.append(r)
        try:
            from openpyxl.chart import BarChart, Reference
            if ws_chartdata.max_row > 1:
                chart = BarChart()
                cats = Reference(ws_chartdata, min_col=1, min_row=2, max_row=ws_chartdata.max_row)
                vals1 = Reference(ws_chartdata, min_col=2, min_row=2, max_row=ws_chartdata.max_row)
                vals2 = Reference(ws_chartdata, min_col=3, min_row=2, max_row=ws_chartdata.max_row)
                chart.add_data(vals1, titles_from_data=False)
                chart.add_data(vals2, titles_from_data=False)
                chart.set_categories(cats)
                chart.title = "Sum DOC and DRR by Brand + Parent"
                ws_chart = wb.create_sheet("Chart")
                ws_chart.add_chart(chart, "A1")
        except Exception:
            pass

    # HowToPivot
    try:
        how = wb.create_sheet("HowToPivot")
        how.append(["How to create interactive PivotTable + Slicer in Excel (no VBA)"])
        how.append([])
        how.append(["1) In Excel: Insert → PivotTable"])
        how.append([f"   - Select the table named 'DataTable' on the sheet: {sheet_name if 'sheet_name' in locals() else 'Data'}"])
        how.append(["   - Place the PivotTable on a new worksheet."])
        how.append([])
        how.append(["2) In PivotField list: drag 'Brand' and '(Parent) ASIN' (or your parent column) into Rows (Brand first)."])
        how.append(["   - Drag 'DOC' and 'DRR' into Values (set aggregation = Sum)."])
        how.append(["   - Drag 'Total CP' into Values (Aggregation = Sum)."])
        how.append([])
        how.append(["3) To add a Slicer: Insert → Slicer → choose 'Brand' and '(Parent) ASIN'."])
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# -------------------------
# Inventory Report Builder
# -------------------------
def build_inventory_report(inventory_df: pd.DataFrame, pm_df: pd.DataFrame) -> pd.DataFrame:
    inv = inventory_df.copy()
    inv.columns = inv.columns.str.strip()
    pm_df.columns = pm_df.columns.str.strip()

    asin_col = [c for c in inv.columns if c.lower() == "asin"][0]
    inv[asin_col] = inv[asin_col].astype(str)

    pm_lookup = (
        pm_df[["ASIN", "Brand", "Brand Manager", "Vendor SKU Codes", "CP"]]
        .drop_duplicates(subset=["ASIN"])
        .copy()
    )

    pm_lookup["ASIN"] = pm_lookup["ASIN"].astype(str)
    pm_lookup["CP"] = pd.to_numeric(pm_lookup["CP"], errors="coerce").fillna(0)

    inv = inv.merge(
        pm_lookup,
        how="left",
        left_on=asin_col,
        right_on="ASIN"
    )

    inv.rename(columns={"Vendor SKU Codes": "Vendor SKU"}, inplace=True)

    wh_col = next((c for c in inv.columns if "afn-warehouse" in c.lower()), None)

    if wh_col:
        inv[wh_col] = pd.to_numeric(inv[wh_col], errors="coerce").fillna(0)
        inv["As per Qty"] = (inv[wh_col] * inv["CP"]).round(2)
    else:
        inv["As per Qty"] = 0

    return inv

# -------------------------
# UI: file upload + processing
# -------------------------
st.title("📊 Inventory Analysis Dashboard")
st.markdown("Upload your files and analyze inventory with Days of Coverage (DOC) calculations")

with st.sidebar:
    st.header("⚙️ Configuration")
    no_of_days = st.number_input(
        "Enter the number of days",
        min_value=1,
        value=31,
        step=1,
        help="This represents the time period for your sales data. DRR = Total Orders÷ Days.",
    )
    st.info(
        """
- DRR (Daily Run Rate) = Total Orders ÷ Days
- DOC (Days of Coverage) = afn-fulfillable-quantity ÷ DRR
"""
    )

st.header("📁 Upload Files")
col1, col2, col3, col4 = st.columns(4)
with col1:
    business_file = st.file_uploader("Upload Business Report CSV/Excel", type=["csv","xlsx"], key="business")
with col2:
    pm_file = st.file_uploader("Upload PM Excel/CSV", type=["xlsx", "csv"], key="pm")
with col3:
    inventory_file = st.file_uploader("Upload Manage Inventory CSV/Excel", type=["csv","xlsx"], key="inventory")
    
with col4:
    inventory_listing_file = st.file_uploader(
        "Upload Inactive Listing Report",
        type=["csv", "xlsx"],
        key="inventory_listing"
    )


st.markdown("---")

if st.button("🚀 Process Data"):
    if (
    business_file is None
    or pm_file is None
    or inventory_file is None
    or inventory_listing_file is None):
        st.error("⚠️ Please upload all four files before processing!")

    elif no_of_days <= 0:
        st.error("⚠️ Number of days must be greater than 0!")
    else:
        with st.spinner("Processing data..."):
            try:
                original = pd.read_csv(business_file)

                # ---------- SAFE SKU COLUMN DETECTION ----------
                sku_col = next(
                    (c for c in original.columns if c.strip().lower() == "sku"),
                    None
                )

                if not sku_col:
                    st.error("❌ Business Report file is invalid. Required column missing: SKU")
                    st.stop()

                # Normalize column name to 'SKU'
                if sku_col != "SKU":
                    original.rename(columns={sku_col: "SKU"}, inplace=True)


                if pm_file.name.endswith(".xlsx"):
                    pm_read= pd.read_excel(pm_file)
                else:
                    pm_read = pd.read_csv(pm_file)

                inventory = pd.read_csv(inventory_file)
                inventory.columns = inventory.columns.str.strip()
                inventory.iloc[:, 0] = inventory.iloc[:, 0].astype(str)
                
                # -------------------------
                # Inventory Report creation
                # -------------------------
                inventory_report_df = build_inventory_report(
                    inventory_df=inventory,
                    pm_df=pm_read
                )

                st.session_state["inventory_report"] = inventory_report_df

                # -------------------------
                # Inventory Listing file
                # -------------------------
                if inventory_listing_file.name.endswith(".xlsx"):
                    inventory_listing = pd.read_excel(inventory_listing_file)
                else:
                    inventory_listing = pd.read_csv(inventory_listing_file)

                inventory_listing.columns = inventory_listing.columns.str.strip()

                # process pm columns C:G
                # -------------------------
                # PM cleanup + merge (SAFE – no duplicates)
                # -------------------------
                pm = pm_read.iloc[:, 2:7].copy()
                pm.columns = ["Amazon Sku Name", "D", "Brand Manager", "F", "Brand"]

                pm["Amazon Sku Name"] = pm["Amazon Sku Name"].astype(str)

                # # CRITICAL: ensure one row per SKU
                # pm = (
                #     pm
                #     .dropna(subset=["Amazon Sku Name"])
                #     .drop_duplicates(subset=["Amazon Sku Name"], keep="first")
                # )

                # SINGLE merge instead of two
                original = original.merge(
                    pm[["Amazon Sku Name", "Brand Manager", "Brand"]],
                    how="left",
                    left_on="SKU",
                    right_on="Amazon Sku Name"
                )

                # reposition columns (optional – your existing logic)
                if "Title" in original.columns:
                    for col_name in ["Brand Manager", "Brand"]:
                        if col_name in original.columns:
                            col = original.pop(col_name)
                            insert_pos = original.columns.get_loc("Title")
                            original.insert(insert_pos, col_name, col)


                # inventory mapping
                # ---------- SAFE afn-fulfillable-quantity detection ----------
                fulfillable_col = next(
                    (c for c in inventory.columns if "afn-fulfillable" in c.lower()),
                    None
                )

                if fulfillable_col:
                    mi_map = inventory.set_index(inventory.columns[0])[fulfillable_col]
                    original["afn-fulfillable-quantity"] = (
                        original["SKU"].map(mi_map).fillna(0)
                    )
                else:
                    original["afn-fulfillable-quantity"] = 0


                reserved_col = next(
                    (c for c in inventory.columns if "afn-reserved" in c.lower()),
                    None
                )

                if reserved_col:
                    mi_res_map = inventory.set_index(inventory.columns[0])[reserved_col]
                    original["afn-reserved-quantity"] = (
                        original["SKU"].map(mi_res_map).fillna(0)
                    )
                else:
                    original["afn-reserved-quantity"] = 0


                # clean & compute DRR/DOC
                # -------------------------
                # Clean & compute Total Orders
                # -------------------------
                for col in ["Total Order Items", "Total Order Items - B2B"]:
                    if col in original.columns:
                        original[col] = (
                            original[col]
                            .astype(str)
                            .str.replace("\u00A0", "", regex=False)
                            .str.replace(",", "", regex=False)
                            .str.replace(r"[^\d\.\-]", "", regex=True)
                        )
                        original[col] = pd.to_numeric(original[col], errors="coerce").fillna(0)
                    else:
                        original[col] = 0

                # ✅ NEW COLUMN
                original["Total Orders"] = (
                    original["Total Order Items"] + original["Total Order Items - B2B"]
                )


                original["DRR"] = (original["Total Orders"] / no_of_days).round(2)
                original["afn-fulfillable-quantity"] = pd.to_numeric(original["afn-fulfillable-quantity"], errors="coerce")
                original["DOC"] = (
                    original["afn-fulfillable-quantity"] /
                    original["DRR"].replace(0, pd.NA)
                ).round(2)

                original["DOC"] = original["DOC"].fillna(0)


                # -------------------------
                # REAL Vendor SKU, CP, Total CP (NO EXCEL FORMULA)
                # -------------------------
                # -------------------------
                # PM lookup (ASIN based)
                # -------------------------
                required_cols = ["ASIN", "Vendor SKU Codes", "CP"]
                missing = [c for c in required_cols if c not in pm_read.columns]
                if missing:
                    raise ValueError(f"PM file missing columns: {missing}")

                # -------------------------
                # PM lookup (DEDUPED by ASIN)
                # -------------------------
                pm_lookup = (
                    pm_read[required_cols]
                    .copy()
                )

                pm_lookup["ASIN"] = pm_lookup["ASIN"].astype(str)
                pm_lookup["CP"] = pd.to_numeric(pm_lookup["CP"], errors="coerce").fillna(0)

                if "(Parent) ASIN" in original.columns:
                    original["(Parent) ASIN"] = original["(Parent) ASIN"].astype(str)

                    original = original.merge(
                        pm_lookup,
                        how="left",
                        left_on="(Parent) ASIN",
                        right_on="ASIN"
                    )

                    original.rename(
                        columns={"Vendor SKU Codes": "Vendor SKU"},
                        inplace=True
                    )

                    original["Total CP"] = (
                        original["afn-fulfillable-quantity"].fillna(0) * original["CP"]
                    ).round(2)

                    original.drop(columns=["ASIN"], inplace=True, errors="ignore")

                else:
                    original["Vendor SKU"] = ""
                    original["CP"] = 0
                    original["Total CP"] = 0
                    
                # ✅ SAFE DEDUPLICATION:
                # same Parent ASIN + same SKU → remove
                # same Parent ASIN + different SKU → keep
                if {"(Parent) ASIN", "SKU"}.issubset(original.columns):
                    original = original.drop_duplicates(
                        subset=["(Parent) ASIN", "SKU"],
                        keep="first"
                    )


                # -------------------------
                # Seller SKU mapping (EXACT Excel VLOOKUP equivalent)
                # -------------------------
                if "SKU" in original.columns and "seller-sku" in inventory_listing.columns:

                    original["SKU"] = original["SKU"].astype(str)
                    inventory_listing["seller-sku"] = inventory_listing["seller-sku"].astype(str)

                    seller_sku_set = set(inventory_listing["seller-sku"].dropna())

                    original["Seller SKU"] = original["SKU"].apply(
                        lambda x: x if x in seller_sku_set else ""
                    )

                else:
                    original["Seller SKU"] = ""
                    st.warning("⚠️ seller-sku column not found in Inventory Listing file")
                
                # -------------------------
                # Listing Status
                # -------------------------
                if {"SKU", "Seller SKU"}.issubset(original.columns):

                    original["Listing Status"] = original.apply(
                        lambda r: "Listing Close" if str(r["SKU"]) == str(r["Seller SKU"]) and r["Seller SKU"] != "" else " ",
                        axis=1
                    )

                else:
                    original["Listing Status"] = ""


                st.success("✅ Data processed successfully!")
                tab_business, tab_inventory = st.tabs(["🧾 Business Report", "📦 Inventory Report"])
                
                # -------------------------
                # SKU vs Total Orders Pivot
                # -------------------------
                pivot_df = (
                    original
                    .groupby("SKU", dropna=False)["Total Orders"]
                    .sum()
                    .reset_index()
                    .sort_values("Total Orders", ascending=False)
                )
                
                # -------------------------
                # Business Sales Qty in Inventory Report (VLOOKUP via map)
                # -------------------------

                # Ensure SKU column exists in inventory report
                inv_sku_col = None
                for c in inventory_report_df.columns:
                    if c.lower() in ["sku"]:
                        inv_sku_col = c
                        break

                # Build lookup dictionary from pivot_df
                sales_lookup = (
                    pivot_df
                    .drop_duplicates(subset=["SKU"])
                    .set_index("SKU")["Total Orders"]
                    .to_dict()
                )

                # Apply VLOOKUP-style mapping
                if inv_sku_col:
                    inventory_report_df[inv_sku_col] = inventory_report_df[inv_sku_col].astype(str)

                    inventory_report_df["Business Sales Qty"] = (
                        inventory_report_df[inv_sku_col]
                        .map(sales_lookup)
                        .fillna(0)
                        .astype(int)
                    )
                else:
                    inventory_report_df["Business Sales Qty"] = 0
                    
                inventory_report_df["DRR"] = (
                    inventory_report_df["Business Sales Qty"] / no_of_days
                ).round(2)

                warehouse_col = next(
                    (c for c in inventory_report_df.columns if "afn-warehouse" in c.lower()),
                    None
                )
                if warehouse_col:
                    inventory_report_df["DOC"] = (
                        inventory_report_df[warehouse_col] /
                        inventory_report_df["DRR"].replace(0, pd.NA)
                    ).fillna(0)
                else:
                    inventory_report_df["DOC"] = 0


                st.session_state["sku_pivot"] = pivot_df
                
                # display metrics
                st.header("📈 Processed Results")
                with tab_business:
                    st.header("📈 Business Report")
                    c1, c2, c3, c4 = st.columns(4)
                    with c1:
                        st.metric("Total Products", len(original))
                    with c2:
                        st.metric("Critical Stock (< 7 days)", int((original["DOC"] < 7).sum()))
                    with c3:
                        st.metric("Average DOC", f"{original['DOC'].mean():.2f} days")
                    with c4:
                        st.metric("Total Orders", f"{original['Total Orders'].sum():,.0f}")

                    st.markdown("---")

                    display_cols = [
                        "(Child) ASIN",
                        "Brand",
                        "Brand Manager",
                        "SKU",
                        "Title",
                        "Units Ordered",
                        "Total Order Items",
                        "Total Order Items - B2B",
                        "Total Orders",
                        "afn-fulfillable-quantity",
                        "afn-reserved-quantity",
                        "DRR",
                        "DOC",
                        "Vendor SKU",
                        "CP",
                        "Total CP",
                        "Seller SKU",
                        "Listing Status"
                    ]
                    display_cols = [col for col in display_cols if col in original.columns]
                    display_df = original[display_cols].copy()
                    styled_df = display_df.style.map(color_doc, subset=["DOC"])

                    st.dataframe(styled_df, use_container_width=True, height=600)


                    st.markdown("---")

                    # Excel export UI (template-first)
                    colA, colB = st.columns(2)
                    with colA:
                        # plain CSV download
                        csv_buf = io.StringIO()
                        original.to_csv(csv_buf, index=False)
                        st.download_button(
                            "📥 Download CSV",
                            data=csv_buf.getvalue(),
                            file_name=f"processed_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv",
                        )

                        # XLSX download that looks like CSV but with DOC conditional formatting
                        xlsx_bytes = create_excel_with_doc_format(original)
                        st.download_button(
                            "📥 Download CSV as Excel (DOC formatted)",
                            data=xlsx_bytes,
                            file_name=f"processed_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}_DOCformatted.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                    with colB:
                        st.markdown("### 📥 Excel Export Center")
                        
                        # 1. Selection settings (Always visible)
                        brands = sorted(original["Brand"].dropna().astype(str).unique().tolist()) if "Brand" in original.columns else []
                        selected_export = st.multiselect("Filter brands for export", options=brands, default=brands, key="exp_brands")
                        
                        # 2. Logic Choice (Using a selectbox ensures the download button stays visible)
                        export_choice = st.selectbox("Choose Export Type", ["-- Select --", "Overstock (DOC >= 90)", "OOS (Qty = 0)"])

                        if export_choice != "-- Select --":
                            sort_desc = True if "Overstock" in export_choice else False
                            
                            # Filter the dataframe based on choice
                            if sort_desc:
                                df_export = filter_overstock(original)
                            else:
                                df_export = filter_oos(original)
                            
                            # Apply brand filters
                            if selected_export:
                                df_export = df_export[df_export["Brand"].isin(selected_export)].copy()
                            
                            df_export = df_export.sort_values(by="DOC", ascending=(not sort_desc)).reset_index(drop=True)

                            # Template Detection logic
                            base_dir = os.path.dirname(__file__)
                            tmpl_xlsm = os.path.join(base_dir, "pivot_template.xlsm")
                            tmpl_xlsx = os.path.join(base_dir, "pivot_template.xlsx")
                            template_path = tmpl_xlsm if os.path.exists(tmpl_xlsm) else (tmpl_xlsx if os.path.exists(tmpl_xlsx) else None)

                            # Generate Workbook Bytes
                            final_bytes = None
                            dl_ext = ".xlsx"
                            dl_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                            if template_path:
                                try:
                                    buf = fill_template_and_get_bytes(template_path, df_export, table_name="DataTable")
                                    final_bytes = buf.getvalue()
                                    if template_path.endswith(".xlsm"):
                                        dl_ext = ".xlsm"
                                        dl_mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
                                    st.success("✅ Pivot template applied.")
                                except Exception:
                                    st.warning("⚠️ Template failed. Falling back to generated workbook.")

                            if final_bytes is None:
                                # Detect parent col for fallback
                                parent_col_export = None
                                for c in original.columns:
                                    cle = c.lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", "")
                                    if "parent" in cle:
                                        parent_col_export = c
                                        break
                                
                                fallback_buf = create_fallback_workbook(df_export, sort_desc, "Export", parent_col=parent_col_export)
                                final_bytes = fallback_buf.getvalue()

                            # 3. Persistent Download Button
                            st.download_button(
                                label=f"📥 Download {export_choice} Workbook",
                                data=final_bytes,
                                file_name=f"{'overstock' if sort_desc else 'oos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{dl_ext}",
                                mime=dl_mime,
                                key="main_export_btn"
                            )

                        # --- Consolidated Pivot Display ---
                        st.markdown("---")
                        st.subheader("📊 SKU vs Total Orders (Pivot)")
                        st.dataframe(pivot_df, use_container_width=True, height=400)
                        
                        # Pivot Download button
                        pivot_xl_bytes = create_excel_with_doc_format(pivot_df)
                        st.download_button(
                            "📥 Download SKU Sales Pivot",
                            data=pivot_xl_bytes,
                            file_name=f"sku_sales_pivot_{datetime.now().strftime('%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="pivot_download_btn"
                        )

                    # persist processed data
                    st.session_state["processed_data"] = original
                
                with tab_inventory:
                    st.header("📦 Inventory Report")

                    inv = inventory_report_df

                    st.metric("Total ASINs", len(inv))
                    st.metric("Total Inventory Value", f"{inv['As per Qty'].sum():,.2f}")

                    st.markdown("---")
                    st.dataframe(inv, use_container_width=True, height=600)

                    st.markdown("---")

                    colI1, colI2 = st.columns(2)

                    # CSV
                    with colI1:
                        csv_buf = io.StringIO()
                        inv.to_csv(csv_buf, index=False)
                        st.download_button(
                            "📥 Download Inventory CSV",
                            data=csv_buf.getvalue(),
                            file_name=f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )

                    # Excel
                    with colI2:
                        xlsx_bytes = create_excel_with_doc_format(inv)
                        st.download_button(
                            "📥 Download Inventory Excel",
                            data=xlsx_bytes,
                            file_name=f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                    st.markdown("---")
                    st.subheader("🚫 Inventory with ZERO Business Sales")

                    # Filter: Business Sales Qty == 0
                    if "Business Sales Qty" in inv.columns:
                        zero_sales_df = inv[
                            pd.to_numeric(inv["Business Sales Qty"], errors="coerce").fillna(0) == 0
                        ].copy()
                    else:
                        zero_sales_df = inv.iloc[0:0].copy()

                    st.metric("Zero Sales SKUs", len(zero_sales_df))

                    st.dataframe(zero_sales_df, use_container_width=True, height=500)

                    xlsx_zero = create_excel_with_doc_format(zero_sales_df)
                    st.download_button(
                        "📥 Download Zero Business Sales Inventory (Excel)",
                        data=xlsx_zero,
                        file_name=f"inventory_zero_sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    st.markdown("---")

                    st.subheader("Inventory OOS / Overstock Export")

                    colO1, colO2 = st.columns(2)

                    with colO1:
                        df_oos = inv[inv.get("afn-warehouse-quantity", 0) == 0]
                        buf_oos = create_excel_with_doc_format(df_oos)

                        st.download_button(
                            label="📥 Download Inventory OOS Excel",
                            data=buf_oos,
                            file_name="inventory_oos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="inv_oos_download"
                        )


                    with colO2:
                        df_over = inv.sort_values("As per Qty", ascending=False)
                        buf_over = create_excel_with_doc_format(df_over)

                        st.download_button(
                            label="📥 Download Inventory Overstock Excel",
                            data=buf_over,
                            file_name="inventory_overstock.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="inv_over_download"
                        )

            except Exception as e:
                st.error("❌ Processing failed.")
                st.exception(e)

# previously processed view
elif "processed_data" in st.session_state:
    orig = st.session_state["processed_data"]
    st.header("📈 Previously Processed Results")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Total Products", len(orig))
    with c2:
        st.metric("Critical Stock (< 7 days)", int((orig["DOC"] < 7).sum()))
    with c3:
        st.metric("Average DOC", f"{orig['DOC'].mean():.2f} days")
    with c4:
        st.metric("Total Orders", f"{orig['Total Orders'].sum():,.0f}")

    st.markdown("---")
    display_cols = [
        "(Child) ASIN",
        "Brand",
        "Brand Manager",
        "SKU",
        "Title",
        "Units Ordered",
        "Total Order Items",
        "Total Order Items - B2B",
        "Total Orders",
        "afn-fulfillable-quantity",
        "afn-reserved-quantity",
        "DRR",
        "DOC",
        "Vendor SKU",
        "CP",
        "Total CP",
        "Seller SKU",
        "Listing Status"
    ]
    display_cols = [col for col in display_cols if col in orig.columns]
    display_df = orig[display_cols].copy()
    styled_df = display_df.style.map(color_doc, subset=["DOC"])
    st.dataframe(styled_df, use_container_width=True, height=600)

    st.markdown("---")
    cA, cB = st.columns(2)
    with cA:
        # plain CSV download
        csv_buf = io.StringIO()
        orig.to_csv(csv_buf, index=False)
        st.download_button(
            "📥 Download CSV",
            data=csv_buf.getvalue(),
            file_name=f"processed_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
        )

        # XLSX download that looks like CSV but with DOC conditional formatting
        xlsx_bytes_prev = create_excel_with_doc_format(orig)
        st.download_button(
            "📥 Download CSV as Excel (DOC formatted)",
            data=xlsx_bytes_prev,
            file_name=f"processed_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}_DOCformatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with cB:
        st.markdown("### Excel export (Template-first pivot)")
        over2 = st.button("📥 Overstock (DOC ↓) - Export Excel (previous)", key="over_prev")
        oos2 = st.button("📥 OOS (DOC ↑) - Export Excel (previous)", key="oos_prev")

        if over2 or oos2:
            sort_desc = True if over2 else False
            parent_col = None
            for c in orig.columns:
                cle = c.lower().replace(" ", "").replace("_", "").replace("(", "").replace(")", "")
                if "parent" in cle:
                    parent_col = c
                    break
            brands = sorted(orig["Brand"].dropna().astype(str).unique().tolist()) if "Brand" in orig.columns else []
            selected = st.multiselect("Filter brands for export (leave empty = all)", options=brands, default=brands)

            # yahan bhi wahi filters apply:
            if over2:
                df_export = filter_overstock(orig)
            else:
                df_export = filter_oos(orig)

            if selected:
                df_export = df_export[df_export["Brand"].isin(selected)].copy()

            df_export = df_export.sort_values(by="DOC", ascending=(not sort_desc)).reset_index(drop=True)

            base_dir = os.path.dirname(__file__)
            tmpl_xlsm = os.path.join(base_dir, "pivot_template.xlsm")
            tmpl_xlsx = os.path.join(base_dir, "pivot_template.xlsx")
            template_path = tmpl_xlsm if os.path.exists(tmpl_xlsm) else (tmpl_xlsx if os.path.exists(tmpl_xlsx) else None)

            final_bytes = None
            template_used = False

            if template_path:
                try:
                    buf = fill_template_and_get_bytes(template_path, df_export, table_name="DataTable")
                    final_bytes = buf.getvalue()
                    template_used = True
                    st.success("✅ Used pivot template — Pivot/Slicer included when opened in Excel.")
                except Exception as te:
                    st.warning("⚠️ Template found but failed to be filled programmatically. Falling back to generated workbook.")
                    st.code(traceback.format_exc())

            if final_bytes is None:
                fallback_buf = create_fallback_workbook(
                    df_export,
                    sort_desc=sort_desc,
                    sheet_name="Overstock" if sort_desc else "OOS",
                    parent_col=parent_col,
                    selected_brands=selected,
                )
                final_bytes = fallback_buf.getvalue()
                st.info("ℹ️ Delivered fallback workbook (DataTable + PivotSummary + ChartData + HowToPivot).")

            # pick correct extension & mime based on template vs fallback
            if template_path and template_used and template_path.lower().endswith(".xlsm"):
                dl_ext = ".xlsm"
                dl_mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
            else:
                dl_ext = ".xlsx"
                dl_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            file_basename = f"{'overstock' if sort_desc else 'oos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            file_name = file_basename + dl_ext

            st.download_button(
                label="Download Excel workbook",
                data=final_bytes,
                file_name=file_name,
                mime=dl_mime,
            )

# footer
st.markdown("---")
st.markdown("<div style='text-align: center; color: #666; padding: 10px;'>Inventory Analysis Dashboard | Built with Streamlit</div>", unsafe_allow_html=True)


